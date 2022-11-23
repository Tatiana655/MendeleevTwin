import numpy as np
from mpmath import *
from twin import *
from intvalpy import *
from operation import *

import matplotlib.pyplot as plt
from openpyxl import load_workbook


def moda(data_int):
    ''' глупый алгоритм для моды на маленьких выборках. Возвращает интервал-значение моды '''
    dt=list()
    for el in data_int:#добавили все точки
        dt.append(el.a)
        dt.append(el.b)
    dt.sort()
    dt_int=list()
    for i in range(len(dt)-1):
        dt_int.append(Interval(dt[i], dt[i+1]))

    count = np.zeros(len(dt_int))

    for i in range(len(dt_int)):
        for j in range(len(data_int)):
            # тут надо написать, если пересечение i и j не пусто, то добавь
            if ((dt_int[i].b > data_int[j].a) and (dt_int[i].a <= data_int[j].a)) or \
                    ((dt_int[i].b >= data_int[j].b) and (dt_int[i].a < data_int[j].b)) or \
                    ((dt_int[i].b == data_int[j].b) and (dt_int[i].a == data_int[j].a)) or\
                    (dt_int[i].b<data_int[j].b and dt_int[i].a>data_int[j].a):

                count[i] += 1
    res = dt_int[np.argmax(count)]
    return res


'''main'''
names=['Hydrogen','Carbon']
workbook = load_workbook(filename="data\\table.xlsx")
outer=list()
inner=list()

for k in [0,1]:
    list_int_data = list()
    sheet = workbook[names[k]]
    '''поиск внешней границы по всем x(1H)'''
    if k==0:
        row_num = np.array([12,13])
    else:
        row_num = np.array([22, 23])
    col_num = [4, 5, 10, 11]

    atomic_low=list()
    atomic_up = list()
    if k==0:
        coef=np.array([1,2])
    else: #carbon
        coef = np.array([12, 13])

    for i in row_num:
        atomic_low.append(coef[0]*sheet.cell(row=i, column=col_num[0]).value +coef[1] *sheet.cell(row=i, column=col_num[1]).value)
        atomic_up.append(coef[0]*sheet.cell(row=i, column=col_num[2]).value +coef[1]* sheet.cell(row=i, column=col_num[3]).value)
        list_int_data.append(Interval(atomic_low[-1],atomic_up[-1]))
    min_bound=min(atomic_low)
    max_bound=max(atomic_up)


    mu=Interval(min_bound,max_bound)
    print("внешняя оценка "+names[k]+" : ", mu)
    #print(list_int_data)
    m=moda(list_int_data)
    print("мода "+names[k]+" : ",m)
    outer.append(mu)
    inner.append(m)

T_h=Twin(inner[0],outer[0])
T_c=Twin(inner[1],outer[1])

T=twins_plus(T_c,Twin(4*T_h.X_l, 4*T_h.X)) #не нашел умножение твина на число, возможно, невнимательный

'''рисование твина. Не нашла готовой функции, обратить внимание на подпись осей абсцисс :('''
print("T_h=", T_h)
print("T_c=", T_c)
print("T=T_c+4*T_h=", T)
fig, ax = plt.subplots()
ax.plot( [T.X_l.a,T.X_l.b],[0,0],c = (0.0, 0.0, 1, 0.5),
        linewidth = 12)

ax.plot( [T.X.a,T.X.b],[0,0],c = (1, 0.0, 0.0, 0.5),
        linewidth = 12)

plt.show()